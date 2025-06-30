import openpyxl
import csv
import os
import logging
from datetime import datetime, date
from project_logger import setup_project_logger
from config import RAW_DATA_DIR, PROCESSED_DATA_DIR

class ExcelProcessor:
    """
    A class to encapsulate the logic for extracting data from Excel files
    and writing it to CSV. Adheres to SOLID principles by separating concerns.
    """
    def __init__(self, raw_data_dir=RAW_DATA_DIR, processed_data_dir=PROCESSED_DATA_DIR):
        """
        Initializes the ExcelProcessor with input and output directories.

        Args:
            raw_data_dir (str): Directory containing raw Excel files.
            processed_data_dir (str): Directory for processed CSV outputs.
        """
        self.raw_data_dir = raw_data_dir
        self.processed_data_dir = processed_data_dir
        self.logger = setup_project_logger("excel_extraction")
        os.makedirs(self.processed_data_dir, exist_ok=True) # Ensure output dir exists

    def _clean_text(self, text):
        """
        Cleans common typographical characters like smart quotes and dashes.
        Replaces them with their ASCII equivalents.
        This remains a utility function, now a private method of the class.
        """
        if text is None:
            return None
        text = str(text)
        text = text.replace('’', "'") # Right single quotation mark
        text = text.replace('‘', "'") # Left single quotation mark
        text = text.replace('“', '"') # Left double quotation mark
        text = text.replace('”', '"') # Right double quotation mark
        text = text.replace('–', '-') # En dash
        text = text.replace('—', '--') # Em dash (longer dash)
        text = text.replace(' ', ' ') # Non-breaking space to regular space
        text = text.replace('\n', ' ') # Remove newlines
        text = text.replace('\r', '') # Remove carriage returns
        text = text.replace('\t', ' ') # Replace tabs with spaces
        text = text.replace('â€', "'") # Common mojibake for apostrophe/right single quote
        text = text.replace('â€œ', '"') # Common mojibake for left double quote
        text = text.replace('â€\uFFFD', '"') # Common mojibake for right double quote
        text = text.replace('â€“', '-') # Common mojibake for en dash
        text = text.replace('â€”', '--') # Common mojibake for em dash
        text = text.replace('…', '...') # Ellipsis to three dots
        # Add more replacements here if other specific characters are found
        return text.strip()

    def _is_cell_highlighted(self, cell):
        """
        Checks if an Excel cell is highlighted (has a non-default fill color).
        This encapsulates the highlighting logic.
        """
        if cell.fill and cell.fill.patternType and cell.fill.fgColor.rgb:
            # openpyxl returns '00000000' for no fill or solid white by default
            # We are looking for any non-transparent fill that's not pure black or pure white
            # (pure black '00000000' is common for default background in some contexts)
            return cell.fill.fgColor.rgb != '00000000' and cell.fill.fgColor.rgb != 'FFFFFFFF'
        return False

    def extract_data_from_single_sheet(self, workbook, sheet_name):
        """
        Extracts article titles, URLs, and source websites from a single sheet
        in an Excel workbook, also identifying if a title cell is highlighted.
        Applies text cleaning to titles and URLs.

        Args:
            workbook (openpyxl.workbook.workbook.Workbook): The loaded openpyxl workbook object.
            sheet_name (str): The name of the sheet to process.

        Returns:
            list: A list of dictionaries, where each dictionary represents an article
                  and contains 'source_website', 'title', 'url', 'is_selected',
                  and 'date' (from sheet name).
        """
        sheet = workbook[sheet_name]
        articles = []
        # Access cells using direct indexing for the first row to get headers
        header_row_values = [sheet.cell(row=1, column=col_idx + 1).value for col_idx in range(sheet.max_column)]

        # Iterate through columns in steps of 2 (A-B, C-D, etc.)
        for col_start_idx in range(0, sheet.max_column, 2):
            source_website = self._clean_text(header_row_values[col_start_idx]) # Clean source website too

            if source_website is None or source_website == '':
                self.logger.debug(f"Skipping empty or null source website column at original Excel column index {col_start_idx + 1}.")
                continue

            # Iterate through rows starting from the second row
            for row_idx in range(2, sheet.max_row + 1):
                title_cell = sheet.cell(row=row_idx, column=col_start_idx + 1)
                url_cell = sheet.cell(row=row_idx, column=col_start_idx + 2)

                title = self._clean_text(title_cell.value)
                url = self._clean_text(url_cell.value)

                # If both title and URL are empty, assume no more articles in this column
                # This prevents processing empty rows at the end of a column block
                if title is None and url is None:
                    self.logger.debug(f"No more data in column for source '{source_website}' from row {row_idx}. Breaking.")
                    break

                is_selected = "Selected" if self._is_cell_highlighted(title_cell) else "Not Selected"
                if is_selected == "Selected":
                     self.logger.debug(f"Title '{title}' for '{source_website}' at row {row_idx} is selected.")


                if title or url: # Only add if there's at least a title or URL
                    articles.append({
                        'source_website': source_website,
                        'title': title,
                        'url': url,
                        'selected': is_selected,
                        'date': sheet_name # Use sheet name as date for the record
                    })
                    self.logger.debug(f"Extracted: '{title}' from '{source_website}' on sheet '{sheet_name}'.")
        return articles

    def process_single_excel_file(self, excel_file_path, output_csv_file_path):
        """
        Processes all sheets in a single Excel file, extracts data, and writes
        the consolidated data to a CSV file.

        Args:
            excel_file_path (str): The path to the input Excel (.xlsx) file.
            output_csv_file_path (str): The path to the output CSV file.
        """
        try:
            workbook = openpyxl.load_workbook(excel_file_path, data_only=True) # data_only=True to get cell values, not formulas
        except FileNotFoundError:
            self.logger.error(f"Error: The Excel file '{excel_file_path}' was not found.")
            return
        except Exception as e:
            self.logger.error(f"Error loading workbook '{excel_file_path}': {e}", exc_info=True)
            return

        all_extracted_data = []
        self.logger.info(f"Processing Excel file: {excel_file_path}")
        self.logger.info(f"Found {len(workbook.sheetnames)} sheets: {', '.join(workbook.sheetnames)}")

        for sheet_name in workbook.sheetnames:
            self.logger.info(f"  Extracting data from sheet: '{sheet_name}'...")
            sheet_data = self.extract_data_from_single_sheet(workbook, sheet_name)
            all_extracted_data.extend(sheet_data)
            self.logger.info(f"    Extracted {len(sheet_data)} entries from '{sheet_name}'.")

        if all_extracted_data:
            csv_fieldnames = ['source_website', 'title', 'url', 'selected', 'date']
            try:
                # Append to CSV if it exists, otherwise create and write header
                file_exists = os.path.isfile(output_csv_file_path)

                with open(output_csv_file_path, 'a', newline='', encoding='utf-8') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=csv_fieldnames)
                    if not file_exists:
                        writer.writeheader()
                    writer.writerows(all_extracted_data)
                self.logger.info(f"\nSuccessfully consolidated data from all sheets and wrote to '{output_csv_file_path}'")
                self.logger.info(f"Total entries written: {len(all_extracted_data)}")
                self.logger.info(f"Total sheets read: {len(workbook.sheetnames)}")
            except Exception as e:
                self.logger.error(f"Error writing consolidated data to CSV file '{output_csv_file_path}': {e}", exc_info=True)
        else:
            self.logger.warning("\nNo data extracted from any sheet to write to CSV.")

    def process_all_excel_files_in_directory(self):
        """
        Processes all Excel files in the designated raw data directory
        and writes the extracted data to corresponding CSV files in the
        processed data directory.
        """
        excel_files = [f for f in os.listdir(self.raw_data_dir) if f.endswith('.xlsx')]

        if not excel_files:
            self.logger.warning(f"No Excel files found in '{self.raw_data_dir}'. Please check the path and file extensions.")
            return False

        today_date = date.today()
        # For simplicity, all data from all Excel files processed today will go into one CSV file
        # You could modify this to create a separate CSV per Excel file if needed.
        output_csv_file_name = f"processed_data_{today_date.strftime('%Y%m%d')}.csv"
        output_csv_file_path = os.path.join(self.processed_data_dir, output_csv_file_name)

        # Remove the CSV file if it exists from a previous run on the same day
        # to prevent duplicate appending if the script is run multiple times
        if os.path.exists(output_csv_file_path):
            self.logger.info(f"Removing existing CSV file: '{output_csv_file_path}' for a fresh run.")
            os.remove(output_csv_file_path)

        for excel_file_name in excel_files:
            excel_file_path = os.path.join(self.raw_data_dir, excel_file_name)
            self.logger.info(f"\n--- Processing '{excel_file_name}' ---")
            self.process_single_excel_file(excel_file_path, output_csv_file_path)
            self.logger.info(f"--- Finished processing '{excel_file_name}' ---\n")

        self.logger.info("All Excel files processed successfully.")
        return True

# --- How to use the code ---
if __name__ == "__main__":
    processor = ExcelProcessor()
    processor.process_all_excel_files_in_directory()